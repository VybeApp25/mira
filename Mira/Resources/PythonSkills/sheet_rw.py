#!/usr/bin/env python3
"""Read or write a spreadsheet (.xlsx, .csv).
Args JSON (stdin):
  read:   {op:"read", path, sheet?, max_rows?}
  write:  {op:"write", path, sheet?, headers?, rows}   # rows = [[val,...],...]
  append: {op:"append", path, sheet?, row}              # row = [val,...]
  csv_to_xlsx: {op:"csv_to_xlsx", csv_path, xlsx_path}
Output JSON (stdout): {success, headers?, rows?, row_count?, path?, error?}
"""
import sys, json, os

def main():
    try:
        args = json.load(sys.stdin)
    except Exception:
        args = {}
    op   = args.get("op", "read")
    path = args.get("path", "")
    ext  = os.path.splitext(path)[1].lower()

    try:
        if op == "read":
            if ext == ".csv":
                import csv
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    rows_raw = list(reader)
                max_rows = args.get("max_rows", 200)
                headers = rows_raw[0] if rows_raw else []
                rows    = rows_raw[1:max_rows+1] if len(rows_raw) > 1 else []
                print(json.dumps({"success": True, "headers": headers, "rows": rows, "row_count": len(rows)}))
            else:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                sheet_name = args.get("sheet") or wb.sheetnames[0]
                ws = wb[sheet_name]
                max_rows = args.get("max_rows", 200)
                all_rows = [[str(c.value) if c.value is not None else "" for c in row] for row in ws.iter_rows()]
                headers = all_rows[0] if all_rows else []
                rows    = all_rows[1:max_rows+1] if len(all_rows) > 1 else []
                print(json.dumps({"success": True, "headers": headers, "rows": rows,
                                  "row_count": len(rows), "sheet": sheet_name}))

        elif op == "write":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = args.get("sheet", "Sheet1")
            if args.get("headers"):
                ws.append(args["headers"])
            for row in args.get("rows", []):
                ws.append(row)
            wb.save(path)
            print(json.dumps({"success": True, "path": path, "row_count": len(args.get("rows", []))}))

        elif op == "append":
            import openpyxl
            wb = openpyxl.load_workbook(path)
            sheet_name = args.get("sheet") or wb.sheetnames[0]
            ws = wb[sheet_name]
            ws.append(args.get("row", []))
            wb.save(path)
            print(json.dumps({"success": True, "path": path}))

        elif op == "csv_to_xlsx":
            import csv, openpyxl
            csv_path  = args.get("csv_path", "")
            xlsx_path = args.get("xlsx_path", "")
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(csv_path, newline="", encoding="utf-8-sig") as f:
                for row in csv.reader(f):
                    ws.append(row)
            wb.save(xlsx_path)
            print(json.dumps({"success": True, "path": xlsx_path}))

        else:
            print(json.dumps({"success": False, "error": f"Unknown op: {op}"}))

    except ImportError as ie:
        pkg = "openpyxl" if "openpyxl" in str(ie) else "unknown"
        print(json.dumps({"success": False, "error": f"{pkg} not installed. Run: pip3 install {pkg}"}))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))

main()
